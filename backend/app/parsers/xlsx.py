"""Excel parser: .xlsx/.xlsm via openpyxl (read-only), legacy .xls via xlrd if importable."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Iterable, Iterator, Optional

from .base import BaseParser, ParsedEvent, clean
from .tabular import ColumnRoles, looks_like_header

XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # OLE2 compound document
XLS_ERROR = "legacy .xls needs the 'xlrd' package (pip install xlrd) - or save the workbook as .xlsx"


def available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.tzinfo is None:
            v = v.replace(tzinfo=timezone.utc)
        return v.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return clean(v)


class XlsxParser(BaseParser):
    name = "Excel workbook"
    family = "document.xlsx"
    binary = True
    extensions = (".xlsx", ".xlsm", ".xls")

    def sniff(self, sample: list[str], filename: str = "", head: bytes = b"") -> float:
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xlsm")):
            return 1.0 if head.startswith(b"PK\x03\x04") else 0.7
        if lower.endswith(".xls"):
            return 1.0 if head.startswith(XLS_MAGIC) else 0.7
        return 0.0

    def parse(self, lines: Iterable[str]) -> Iterator[ParsedEvent]:
        return iter(())

    def parse_bytes(self, data: bytes) -> Iterator[ParsedEvent]:
        if data.startswith(XLS_MAGIC):
            yield from self._parse_xls(data)
            return
        try:
            import openpyxl
        except Exception as exc:
            raise RuntimeError(f"openpyxl not installed: {exc}")
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                yield from self._sheet(ws.title, ws.iter_rows(values_only=True))
        finally:
            wb.close()

    def _parse_xls(self, data: bytes) -> Iterator[ParsedEvent]:
        try:
            import xlrd
        except Exception:
            raise RuntimeError(XLS_ERROR)
        book = xlrd.open_workbook(file_contents=data)
        for sheet in book.sheets():
            def rows(sheet=sheet) -> Iterator[tuple[Any, ...]]:
                for r in range(sheet.nrows):
                    out = []
                    for c in range(sheet.ncols):
                        cell = sheet.cell(r, c)
                        v: Any = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            try:
                                v = xlrd.xldate_as_datetime(v, book.datemode)
                            except Exception:
                                pass
                        out.append(v)
                    yield tuple(out)
            yield from self._sheet(sheet.name, rows())

    def _sheet(self, title: str, rows: Iterable[tuple[Any, ...]]) -> Iterator[ParsedEvent]:
        it = iter(rows)
        first: Optional[list[str]] = None
        second: Optional[list[str]] = None
        row_no = 0
        first_no = 0
        for r in it:  # skip leading empty rows
            row_no += 1
            cells = [_cell(v) for v in r]
            if any(cells):
                first, first_no = cells, row_no
                break
        if first is None:
            return
        pending: list[tuple[int, list[str]]] = []
        for r in it:
            row_no += 1
            second = [_cell(v) for v in r]
            pending.append((row_no, second))
            break
        if looks_like_header(first, second):
            header = [h or f"col{i + 1}" for i, h in enumerate(first)]
        else:
            header = [f"col{i + 1}" for i in range(len(first))]
            pending.insert(0, (first_no, first))
        roles = ColumnRoles(header)

        def emit(no: int, cells: list[str]) -> Optional[ParsedEvent]:
            if not any(cells):
                return None
            raw = "\t".join(cells).rstrip("\t")
            return roles.event(raw, cells, {"sheet": title, "row": str(no)})

        for no, cells in pending:
            ev = emit(no, cells)
            if ev:
                yield ev
        for r in it:
            row_no += 1
            ev = emit(row_no, [_cell(v) for v in r])
            if ev:
                yield ev
